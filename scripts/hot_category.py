# -*- coding: utf-8 -*-

from argparse import ArgumentParser
from collections import defaultdict
from contextlib import closing
from datetime import date, timedelta
from itertools import izip_longest
from locale import LC_ALL, format, setlocale
from sys import modules
from unicodedata import normalize

from celery import Celery
from nltk.tokenize import word_tokenize
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Style
from sqlalchemy.orm import backref, relationship
from ujson import dumps, loads

from keyword_analyzer import get_contents
from keyword_suggester import get_results
from suggested_keywords import get_suggested_keywords
from top_100_explorer import book, category, trend
from utilities import (
    base,
    get_date,
    get_mysql_session,
    get_words,
    stopwords,
    variables,
)

if 'threading' in modules:
    del modules['threading']

setlocale(LC_ALL, 'en_US.UTF-8')

celery = Celery('hot_category')
celery.conf.update(
    BROKER=variables['broker'],
    BROKER_POOL_LIMIT=0,
    CELERY_ACCEPT_CONTENT=['json'],
    CELERY_ACKS_LATE=True,
    CELERY_IGNORE_RESULT=True,
    CELERY_RESULT_SERIALIZER='json',
    CELERY_TASK_SERIALIZER='json',
    CELERYD_LOG_FORMAT='[%(asctime)s: %(levelname)s] %(message)s',
    CELERYD_POOL_RESTARTS=True,
    CELERYD_PREFETCH_MULTIPLIER=1,
    CELERYD_TASK_SOFT_TIME_LIMIT=3600,
    CELERYD_TASK_TIME_LIMIT=7200,
)


class step_1_word(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_1_words'


class step_2_keyword(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_2_keywords'


class step_3_suggested_keyword(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_3_suggested_keywords'


class step_4_word(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_4_words'


class step_5_keyword(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_5_keywords'


class step_6_suggested_keyword(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_6_suggested_keywords'


class step_7_group(base):
    __table_args__ = {
        'autoload': True,
    }
    __tablename__ = 'apps_hot_category_step_7_groups'

    suggested_keywords = relationship(
        'step_6_suggested_keyword',
        backref=backref('groups', lazy='dynamic'),
        lazy='dynamic',
        secondary='apps_hot_category_step_7_groups_suggested_keywords',
    )


class step_7_group_suggested_keyword(base):
    __tablename__ = 'apps_hot_category_step_7_groups_suggested_keywords'
    __table_args__ = {
        'autoload': True,
    }

    group = relationship('step_7_group')
    suggested_keyword = relationship('step_6_suggested_keyword')


def get_titles(category_id, print_length):
    titles = []
    with closing(get_mysql_session()()) as session:
        query = session.query(book.title)
        if print_length == 'short':
            query = query.filter(book.print_length < 100)
        if print_length == 'long':
            query = query.filter(book.print_length >= 100)
        if category_id:
            query = query.join(
                trend,
            ).filter(
                trend.category_id == category_id,
                trend.date == date.today() - timedelta(days=1),
            )
        for b in query.order_by(
            'apps_top_100_explorer_books.id ASC',
        ).execution_options(
            stream_results=True,
        ):
            string = normalize(
                'NFKD', b.title.lower()
            ).encode('ascii', 'ignore').strip()
            if string:
                titles.append(string)
        if category_id:
            for c in session.query(
                category.id,
            ).filter(
                category.category_id == category_id,
            ).order_by(
                'id ASC',
            ).execution_options(
                stream_results=True,
            ):
                for string in get_titles(c.id, print_length):
                    titles.append(string)
    return titles


def step_1(date, category_id, print_length):
    print 'Step 1'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_1_word,
        ).filter(
            step_1_word.category_id == category_id,
            step_1_word.date == date,
            step_1_word.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
        for string, _ in get_words([
            title
            for title in get_titles(category_id, print_length)
        ], 20):
            session.add(step_1_word(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
                'string': string,
            }))
        session.commit()
    print 'Done'


def step_2_reset(date, category_id, print_length):
    print 'Step 2 :: Reset'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_2_keyword,
        ).filter(
            step_2_keyword.category_id == category_id,
            step_2_keyword.date == date,
            step_2_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
    print 'Done'


def step_2_queue(date, category_id, print_length):
    print 'Step 2 :: Queue'
    with closing(get_mysql_session()()) as session:
        strings = [
            word.string
            for word in session.query(
                step_1_word,
            ).filter(
                step_1_word.category_id == category_id,
                step_1_word.date == date,
                step_1_word.print_length == print_length,
            ).order_by(
                'id ASC',
            ).execution_options(
                stream_results=True,
            )
        ]
        for string in strings:
            step_2_process.delay(
                date.isoformat(), category_id, print_length, string, strings,
            )
    print 'Done'


@celery.task
def step_2_process(date, category_id, print_length, string, strings):
    ss = get_results(string, 'com', 'digital-text')
    with closing(get_mysql_session()()) as session:
        for s in ss:
            if session.query(
                step_2_keyword,
            ).filter(
                step_2_keyword.category_id == category_id,
                step_2_keyword.date == date,
                step_2_keyword.print_length == print_length,
                step_2_keyword.string == s,
            ).count():
                continue
            if len(set(word_tokenize(s)).intersection(set(strings))) < 1:
                continue
            session.add(step_2_keyword(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
                'string': s,
            }))
        session.commit()


def step_3_1_reset(date, category_id, print_length):
    print 'Step 3.1 :: Reset'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_3_suggested_keyword,
        ).filter(
            step_3_suggested_keyword.category_id == category_id,
            step_3_suggested_keyword.date == date,
            step_3_suggested_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
    print 'Done'


def step_3_1_queue(date, category_id, print_length):
    print 'Step 3.1 :: Queue'
    with closing(get_mysql_session()()) as session:
        for keyword in session.query(
            step_2_keyword,
        ).filter(
            step_2_keyword.category_id == category_id,
            step_2_keyword.date == date,
            step_2_keyword.print_length == print_length,
        ).order_by(
            'id ASC',
        ).execution_options(
            stream_results=True,
        ):
            step_3_1_process.delay(
                date.isoformat(), category_id, print_length, keyword.string,
            )
    print 'Done'


@celery.task
def step_3_1_process(date, category_id, print_length, string):
    ss = get_suggested_keywords(word_tokenize(string))
    with closing(get_mysql_session()()) as session:
        for s in ss:
            if session.query(
                step_3_suggested_keyword,
            ).filter(
                step_3_suggested_keyword.category_id == category_id,
                step_3_suggested_keyword.date == date,
                step_3_suggested_keyword.print_length == print_length,
                step_3_suggested_keyword.string == s,
            ).count():
                continue
            session.add(step_3_suggested_keyword(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
                'string': s,
            }))
        session.commit()


def step_3_2_reset(date, category_id, print_length):
    print 'Step 3.2 :: Reset'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_3_suggested_keyword,
        ).filter(
            step_3_suggested_keyword.category_id == category_id,
            step_3_suggested_keyword.date == date,
            step_3_suggested_keyword.print_length == print_length,
        ).update(
            {
                'average_price': 0.00,
                'average_print_length': 0.00,
                'buyer_behavior': '',
                'competition': '',
                'count': 0,
                'optimization': '',
                'popularity': '',
                'score': 0.00,
                'spend': 0.00,
                'items': '[]',
                'words': '[]',
            },
            synchronize_session=False,
        )
        session.commit()
    print 'Done'


def step_3_2_queue(date, category_id, print_length):
    print 'Step 3.2 :: Queue'
    with closing(get_mysql_session()()) as session:
        for suggested_keyword in session.query(
            step_3_suggested_keyword,
        ).filter(
            step_3_suggested_keyword.category_id == category_id,
            step_3_suggested_keyword.date == date,
            step_3_suggested_keyword.print_length == print_length,
            step_3_suggested_keyword.score == 0.00,
        ).order_by(
            'id ASC',
        ).execution_options(
            stream_results=True,
        ):
            step_3_2_process.delay(
                category_id,
                print_length,
                suggested_keyword.id,
                suggested_keyword.string,
            )
    print 'Done'


@celery.task
def step_3_2_process(category_id, print_length, id, string):
    contents = get_contents(string, 'com')
    if not contents:
        return
    with closing(get_mysql_session()()) as session:
        suggested_keyword = session.query(step_3_suggested_keyword).get(id)
        if not suggested_keyword:
            return
        suggested_keyword.count = contents['count'][0]
        suggested_keyword.buyer_behavior = contents['buyer_behavior'][1]
        suggested_keyword.competition = contents['competition'][1]
        suggested_keyword.optimization = contents['optimization'][1]
        suggested_keyword.popularity = contents['popularity'][1]
        suggested_keyword.spend = contents['spend'][0][0]
        suggested_keyword.average_price = contents['average_price'][0]
        suggested_keyword.average_print_length = contents['average_length'][0]
        suggested_keyword.score = contents['score'][0]
        suggested_keyword.items = dumps(contents['items'])
        suggested_keyword.words = dumps(contents['words'])
        session.add(suggested_keyword)
        session.commit()


def step_4(date, category_id, print_length):
    print 'Step 4'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_4_word,
        ).filter(
            step_4_word.category_id == category_id,
            step_4_word.date == date,
            step_4_word.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
        for string, _ in get_words([
            word[0]
            for suggested_keyword in session.query(
                step_3_suggested_keyword,
            ).filter(
                step_3_suggested_keyword.category_id == category_id,
                step_3_suggested_keyword.date == date,
                step_3_suggested_keyword.print_length == print_length,
                step_3_suggested_keyword.score >= 50.00,
            ).order_by(
                'id ASC',
            ).execution_options(
                stream_results=True,
            )
            for word in loads(suggested_keyword.words)
        ], 20):
            session.add(step_4_word(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
                'string': string,
            }))
        session.commit()
    print 'Done'


def step_5_reset(date, category_id, print_length):
    print 'Step 5 :: Reset'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_5_keyword,
        ).filter(
            step_5_keyword.category_id == category_id,
            step_5_keyword.date == date,
            step_5_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
    print 'Done'


def step_5_queue(date, category_id, print_length):
    print 'Step 5 :: Queue'
    with closing(get_mysql_session()()) as session:
        strings = [
            word.string
            for word in session.query(
                step_4_word,
            ).filter(
                step_4_word.category_id == category_id,
                step_4_word.date == date,
                step_4_word.print_length == print_length,
            ).order_by(
                'id ASC',
            ).execution_options(
                stream_results=True,
            )
        ]
        for string in strings:
            step_5_process.delay(
                date.isoformat(), category_id, print_length, string, strings,
            )
    print 'Done'


@celery.task
def step_5_process(date, category_id, print_length, string, strings):
    ss = get_results(string, 'com', 'digital-text')
    with closing(get_mysql_session()()) as session:
        for s in ss:
            if session.query(
                step_5_keyword,
            ).filter(
                step_5_keyword.category_id == category_id,
                step_5_keyword.date == date,
                step_5_keyword.print_length == print_length,
                step_5_keyword.string == s,
            ).count():
                continue
            if len(set(word_tokenize(s)).intersection(set(strings))) < 1:
                continue
            session.add(step_5_keyword(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
                'string': s,
            }))
        session.commit()


def step_6_1_reset(date, category_id, print_length):
    print 'Step 6.1 :: Reset'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_6_suggested_keyword,
        ).filter(
            step_6_suggested_keyword.category_id == category_id,
            step_6_suggested_keyword.date == date,
            step_6_suggested_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
    print 'Done'


def step_6_1_queue(date, category_id, print_length):
    print 'Step 6.1 :: Queue'
    with closing(get_mysql_session()()) as session:
        for keyword in session.query(
            step_5_keyword,
        ).filter(
            step_5_keyword.category_id == category_id,
            step_5_keyword.date == date,
            step_5_keyword.print_length == print_length,
        ).order_by(
            'id ASC',
        ).execution_options(
            stream_results=True,
        ):
            step_6_1_process.delay(
                date.isoformat(), category_id, print_length, keyword.string,
            )
    print 'Done'


@celery.task
def step_6_1_process(date, category_id, print_length, string):
    ss = get_suggested_keywords(word_tokenize(string))
    with closing(get_mysql_session()()) as session:
        for s in ss:
            if session.query(
                step_6_suggested_keyword,
            ).filter(
                step_6_suggested_keyword.category_id == category_id,
                step_6_suggested_keyword.date == date,
                step_6_suggested_keyword.print_length == print_length,
                step_6_suggested_keyword.string == s,
            ).count():
                continue
            session.add(step_6_suggested_keyword(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
                'string': s,
            }))
        session.commit()


def step_6_2_reset(date, category_id, print_length):
    print 'Step 6.2 :: Reset'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_6_suggested_keyword,
        ).filter(
            step_6_suggested_keyword.category_id == category_id,
            step_6_suggested_keyword.date == date,
            step_6_suggested_keyword.print_length == print_length,
        ).update(
            {
                'average_price': 0.00,
                'average_print_length': 0.00,
                'buyer_behavior': '',
                'competition': '',
                'count': 0,
                'optimization': '',
                'popularity': '',
                'score': 0.00,
                'spend': 0.00,
                'items': '[]',
                'words': '[]',
            },
            synchronize_session=False,
        )
        session.commit()
    print 'Done'


def step_6_2_queue(date, category_id, print_length):
    print 'Step 6.2 :: Queue'
    with closing(get_mysql_session()()) as session:
        for suggested_keyword in session.query(
            step_6_suggested_keyword,
        ).filter(
            step_6_suggested_keyword.category_id == category_id,
            step_6_suggested_keyword.date == date,
            step_6_suggested_keyword.print_length == print_length,
            step_6_suggested_keyword.score == 0.00,
        ).order_by(
            'id ASC',
        ).execution_options(
            stream_results=True,
        ):
            step_6_2_process.delay(
                category_id,
                print_length,
                suggested_keyword.id,
                suggested_keyword.string,
            )
    print 'Done'


@celery.task
def step_6_2_process(category_id, print_length, id, string):
    contents = get_contents(string, 'com')
    if not contents:
        return
    with closing(get_mysql_session()()) as session:
        suggested_keyword = session.query(step_6_suggested_keyword).get(id)
        if not suggested_keyword:
            return
        suggested_keyword.count = contents['count'][0]
        suggested_keyword.buyer_behavior = contents['buyer_behavior'][1]
        suggested_keyword.competition = contents['competition'][1]
        suggested_keyword.optimization = contents['optimization'][1]
        suggested_keyword.popularity = contents['popularity'][1]
        suggested_keyword.spend = contents['spend'][0][0]
        suggested_keyword.average_price = contents['average_price'][0]
        suggested_keyword.average_print_length = contents['average_length'][0]
        suggested_keyword.score = contents['score'][0]
        suggested_keyword.items = dumps(contents['items'])
        suggested_keyword.words = dumps(contents['words'])
        session.add(suggested_keyword)
        session.commit()


def step_7(date, category_id, print_length):
    print 'Step 7'
    with closing(get_mysql_session()()) as session:
        session.query(
            step_7_group,
        ).filter(
            step_7_group.category_id == category_id,
            step_7_group.date == date,
            step_7_group.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.commit()
    with closing(get_mysql_session()()) as session:
        items = defaultdict(set)
        suggested_keywords = session.query(
            step_6_suggested_keyword,
        ).filter(
            step_6_suggested_keyword.category_id == category_id,
            step_6_suggested_keyword.date == date,
            step_6_suggested_keyword.print_length == print_length,
            step_6_suggested_keyword.score >= 50.00,
        ).order_by(
            'score DESC',
        ).execution_options(
            stream_results=True,
        )[:]
        range_ = range(len(suggested_keywords))
        for outer in range_:
            items[outer].add(suggested_keywords[outer])
            for inner in range(outer + 1, len(range_)):
                if suggested_keywords[inner].string in [
                    suggested_keyword_.string
                    for value in items.values()
                    for suggested_keyword_ in value
                ]:
                    continue
                if len(
                    set(
                        word
                        for word in word_tokenize(
                            suggested_keywords[outer].string
                        )
                        if word not in stopwords
                    ).intersection(
                        word
                        for word in word_tokenize(
                            suggested_keywords[inner].string
                        )
                        if word not in stopwords
                    )
                ) < 2:
                    continue
                items[outer].add(suggested_keywords[inner])
        for value in items.values():
            group = step_7_group(**{
                'category_id': category_id,
                'date': date,
                'print_length': print_length,
            })
            for suggested_keyword in value:
                session.add(step_7_group_suggested_keyword(**{
                    'date': date,
                    'group': group,
                    'suggested_keyword': suggested_keyword,
                }))
        session.commit()
    print 'Done'


def xlsx(date, category_id, print_length):
    print 'XLSX'
    th_center = Style(
        alignment=Alignment(
            horizontal='center',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=True,
            italic=False,
            name='Calibri',
            size=12,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    th_left = Style(
        alignment=Alignment(
            horizontal='left',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=True,
            italic=False,
            name='Calibri',
            size=12,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    th_right = Style(
        alignment=Alignment(
            horizontal='right',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=True,
            italic=False,
            name='Calibri',
            size=12,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    td_center = Style(
        alignment=Alignment(
            horizontal='center',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=False,
            italic=False,
            name='Calibri',
            size=10,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    td_left = Style(
        alignment=Alignment(
            horizontal='left',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=False,
            italic=False,
            name='Calibri',
            size=10,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )
    td_right = Style(
        alignment=Alignment(
            horizontal='right',
            indent=0,
            shrink_to_fit=False,
            text_rotation=0,
            vertical='center',
            wrap_text=False,
        ),
        font=Font(
            bold=False,
            italic=False,
            name='Calibri',
            size=10,
            strike=False,
            underline='none',
            vertAlign=None,
        ),
    )

    workbook = Workbook()

    for worksheet in workbook.worksheets:
        workbook.remove_sheet(worksheet)

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 1 - Words'
    worksheet.cell('A1').style = th_left
    worksheet.cell('A1').value = 'String'
    with closing(get_mysql_session()()) as session:
        row = 1
        for word in session.query(
            step_1_word,
        ).filter(
            step_1_word.category_id == category_id,
            step_1_word.date == date,
            step_1_word.print_length == print_length,
        ).order_by(
            'string ASC',
        ).execution_options(
            stream_results=True,
        ):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = word.string
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = td_left
    worksheet.column_dimensions['A'].width = 50

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 2 - Keywords'
    worksheet.cell('A1').style = th_left
    worksheet.cell('A1').value = 'String'
    with closing(get_mysql_session()()) as session:
        row = 1
        for keyword in session.query(
            step_2_keyword,
        ).filter(
            step_2_keyword.category_id == category_id,
            step_2_keyword.date == date,
            step_2_keyword.print_length == print_length,
        ).order_by(
            'string ASC',
        ).execution_options(
            stream_results=True,
        ):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = keyword.string
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = td_left
    worksheet.column_dimensions['A'].width = 50

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 3 - Suggested Keywords'
    worksheet.cell('A1').style = th_left
    worksheet.cell('A1').value = 'String'
    worksheet.cell('B1').style = th_right
    worksheet.cell('B1').value = 'Count'
    worksheet.cell('C1').style = th_center
    worksheet.cell('C1').value = 'Buyer Behavior'
    worksheet.cell('D1').style = th_center
    worksheet.cell('D1').value = 'Competition'
    worksheet.cell('E1').style = th_center
    worksheet.cell('E1').value = 'Optimization'
    worksheet.cell('F1').style = th_center
    worksheet.cell('F1').value = 'Popularity'
    worksheet.cell('G1').style = th_right
    worksheet.cell('G1').value = 'Spend'
    worksheet.cell('H1').style = th_right
    worksheet.cell('H1').value = 'Average Price'
    worksheet.cell('I1').style = th_right
    worksheet.cell('I1').value = 'Average Print Length'
    worksheet.cell('J1').style = th_right
    worksheet.cell('J1').value = 'Score'
    with closing(get_mysql_session()()) as session:
        row = 1
        for suggested_keyword in session.query(
            step_3_suggested_keyword,
        ).filter(
            step_3_suggested_keyword.category_id == category_id,
            step_3_suggested_keyword.date == date,
            step_3_suggested_keyword.print_length == print_length,
        ).order_by(
            'score DESC',
        ).execution_options(
            stream_results=True,
        ):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = td_left
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = suggested_keyword.string
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).value = format('%d', suggested_keyword.count, grouping=True)
            worksheet.cell('C%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('C%(row)s' % {
                'row': row,
            }).value = suggested_keyword.buyer_behavior
            worksheet.cell('D%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('D%(row)s' % {
                'row': row,
            }).value = suggested_keyword.competition
            worksheet.cell('E%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('E%(row)s' % {
                'row': row,
            }).value = suggested_keyword.optimization
            worksheet.cell('F%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('F%(row)s' % {
                'row': row,
            }).value = suggested_keyword.popularity
            worksheet.cell('G%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('G%(row)s' % {
                'row': row,
            }).value = format('%.2f', suggested_keyword.spend, grouping=True)
            worksheet.cell('H%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('H%(row)s' % {
                'row': row,
            }).value = format(
                '%.2f', suggested_keyword.average_price, grouping=True,
            )
            worksheet.cell('I%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('I%(row)s' % {
                'row': row,
            }).value = format(
                '%.2f', suggested_keyword.average_print_length, grouping=True,
            )
            worksheet.cell('J%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('J%(row)s' % {
                'row': row,
            }).value = format('%.2f', suggested_keyword.score, grouping=True)
    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 15

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 4 - Words'
    worksheet.cell('A1').style = th_left
    worksheet.cell('A1').value = 'String'
    with closing(get_mysql_session()()) as session:
        row = 1
        for word in session.query(
            step_4_word,
        ).filter(
            step_4_word.category_id == category_id,
            step_4_word.date == date,
            step_4_word.print_length == print_length,
        ).order_by(
            'string ASC',
        ).execution_options(
            stream_results=True,
        ):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = word.string
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = td_left
    worksheet.column_dimensions['A'].width = 50

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 5 - Keywords'
    worksheet.cell('A1').style = th_left
    worksheet.cell('A1').value = 'String'
    with closing(get_mysql_session()()) as session:
        row = 1
        for keyword in session.query(
            step_5_keyword,
        ).filter(
            step_5_keyword.category_id == category_id,
            step_5_keyword.date == date,
            step_5_keyword.print_length == print_length,
        ).order_by(
            'string ASC',
        ).execution_options(
            stream_results=True,
        ):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = keyword.string
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = td_left
    worksheet.column_dimensions['A'].width = 50

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 6 - Suggested Keywords'
    worksheet.cell('A1').style = th_left
    worksheet.cell('A1').value = 'String'
    worksheet.cell('B1').style = th_right
    worksheet.cell('B1').value = 'Count'
    worksheet.cell('C1').style = th_center
    worksheet.cell('C1').value = 'Buyer Behavior'
    worksheet.cell('D1').style = th_center
    worksheet.cell('D1').value = 'Competition'
    worksheet.cell('E1').style = th_center
    worksheet.cell('E1').value = 'Optimization'
    worksheet.cell('F1').style = th_center
    worksheet.cell('F1').value = 'Popularity'
    worksheet.cell('G1').style = th_right
    worksheet.cell('G1').value = 'Spend'
    worksheet.cell('H1').style = th_right
    worksheet.cell('H1').value = 'Average Price'
    worksheet.cell('I1').style = th_right
    worksheet.cell('I1').value = 'Average Print Length'
    worksheet.cell('J1').style = th_right
    worksheet.cell('J1').value = 'Score'
    with closing(get_mysql_session()()) as session:
        row = 1
        for suggested_keyword in session.query(
            step_6_suggested_keyword,
        ).filter(
            step_6_suggested_keyword.category_id == category_id,
            step_6_suggested_keyword.date == date,
            step_6_suggested_keyword.print_length == print_length,
        ).order_by(
            'score DESC',
        ).execution_options(
            stream_results=True,
        ):
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = td_left
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = suggested_keyword.string
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).value = format('%d', suggested_keyword.count, grouping=True)
            worksheet.cell('C%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('C%(row)s' % {
                'row': row,
            }).value = suggested_keyword.buyer_behavior
            worksheet.cell('D%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('D%(row)s' % {
                'row': row,
            }).value = suggested_keyword.competition
            worksheet.cell('E%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('E%(row)s' % {
                'row': row,
            }).value = suggested_keyword.optimization
            worksheet.cell('F%(row)s' % {
                'row': row,
            }).style = td_center
            worksheet.cell('F%(row)s' % {
                'row': row,
            }).value = suggested_keyword.popularity
            worksheet.cell('G%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('G%(row)s' % {
                'row': row,
            }).value = format('%.2f', suggested_keyword.spend, grouping=True)
            worksheet.cell('H%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('H%(row)s' % {
                'row': row,
            }).value = format(
                '%.2f', suggested_keyword.average_price, grouping=True,
            )
            worksheet.cell('I%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('I%(row)s' % {
                'row': row,
            }).value = format(
                '%.2f', suggested_keyword.average_print_length, grouping=True,
            )
            worksheet.cell('J%(row)s' % {
                'row': row,
            }).style = td_right
            worksheet.cell('J%(row)s' % {
                'row': row,
            }).value = format('%.2f', suggested_keyword.score, grouping=True)
    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 15

    worksheet = workbook.create_sheet()
    worksheet.title = 'Step 7 - Groups'
    with closing(get_mysql_session()()) as session:
        number = 0
        row = 0
        for group in session.query(
            step_7_group,
        ).filter(
            step_7_group.category_id == category_id,
            step_7_group.date == date,
            step_7_group.print_length == print_length,
        ).order_by(
            'id ASC',
        ).execution_options(
            stream_results=True,
        ):
            if not group.suggested_keywords.count() > 1:
                continue
            number += 1
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = th_left
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = 'Group %(number)s' % {
                'number': number,
            }
            worksheet.merge_cells('A%(row)s:B%(row)s' % {
                'row': row,
            })
            row += 1
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).style = Style(
                alignment=Alignment(
                    horizontal='left',
                    indent=0,
                    shrink_to_fit=False,
                    text_rotation=0,
                    vertical='center',
                    wrap_text=False,
                ),
                font=Font(
                    bold=True,
                    italic=False,
                    name='Calibri',
                    size=10,
                    strike=False,
                    underline='none',
                    vertAlign=None,
                ),
            )
            worksheet.cell('A%(row)s' % {
                'row': row,
            }).value = 'Suggested Keywords'
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).style = Style(
                alignment=Alignment(
                    horizontal='left',
                    indent=0,
                    shrink_to_fit=False,
                    text_rotation=0,
                    vertical='center',
                    wrap_text=False,
                ),
                font=Font(
                    bold=True,
                    italic=False,
                    name='Calibri',
                    size=10,
                    strike=False,
                    underline='none',
                    vertAlign=None,
                ),
            )
            worksheet.cell('B%(row)s' % {
                'row': row,
            }).value = 'Books'
            suggested_keywords = [
                (suggested_keyword.string, suggested_keyword.items, )
                for suggested_keyword in group.suggested_keywords.order_by(
                    'apps_hot_category_step_6_suggested_keywords.score DESC',
                ).order_by(
                    'apps_hot_category_step_6_suggested_keywords.score DESC',
                ).execution_options(
                    stream_results=True,
                )
            ]
            titles = [
                item['title'][0]
                for suggested_keyword in suggested_keywords[1:]
                for item in loads(suggested_keyword[1])
            ]
            for iterable in izip_longest(
                [
                    suggested_keyword[0]
                    for suggested_keyword in suggested_keywords
                ],
                [
                    (item['title'][0], item['url'], )
                    for item in loads(suggested_keywords[0][1])
                    if item['title'][0] in titles
                ]
            ):
                row += 1
                worksheet.cell('A%(row)s' % {
                    'row': row,
                }).style = td_left
                worksheet.cell('A%(row)s' % {
                    'row': row,
                }).value = iterable[0]
                if iterable[1]:
                    worksheet.cell('B%(row)s' % {
                        'row': row,
                    }).style = td_left
                    worksheet.cell('B%(row)s' % {
                        'row': row,
                    }).hyperlink = iterable[1][1]
                    worksheet.cell('B%(row)s' % {
                        'row': row,
                    }).value = iterable[1][0]
    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 50

    workbook.save(
        '../tmp/%(date)s-%(category_id)s-%(print_length)s-report.xlsx' % {
            'category_id': arguments.category_id,
            'date': date.isoformat(),
            'print_length': arguments.print_length,
        }
    )
    print 'Done'


def reset(date, category_id, print_length):
    print 'Reset'
    date = date - timedelta(weeks=3)
    with closing(get_mysql_session()()) as session:
        session.query(
            step_1_word,
        ).filter(
            step_1_word.category_id == category_id,
            step_1_word.date < date,
            step_1_word.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_2_keyword,
        ).filter(
            step_2_keyword.category_id == category_id,
            step_2_keyword.date < date,
            step_2_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_3_suggested_keyword,
        ).filter(
            step_3_suggested_keyword.category_id == category_id,
            step_3_suggested_keyword.date < date,
            step_3_suggested_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_4_word,
        ).filter(
            step_4_word.category_id == category_id,
            step_4_word.date < date,
            step_4_word.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_5_keyword,
        ).filter(
            step_5_keyword.category_id == category_id,
            step_5_keyword.date < date,
            step_5_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_6_suggested_keyword,
        ).filter(
            step_6_suggested_keyword.category_id == category_id,
            step_6_suggested_keyword.date < date,
            step_6_suggested_keyword.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_7_group,
        ).filter(
            step_7_group.category_id == category_id,
            step_7_group.date < date,
            step_7_group.print_length == print_length,
        ).delete(
            synchronize_session=False,
        )
        session.query(
            step_7_group_suggested_keyword,
        ).filter(
            step_7_group_suggested_keyword.date < date
        ).delete(
            synchronize_session=False,
        )
        session.commit()
    print 'Done'

if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument(
        '--category-id', dest='category_id', required=True, type=int,
    )
    parser.add_argument('--print-length', dest='print_length', required=True)
    parser.add_argument('--def', dest='def_', required=True)
    parser.set_defaults(reset=False)

    arguments = parser.parse_args()

    locals()[arguments.def_](
        get_date(), arguments.category_id, arguments.print_length
    )
